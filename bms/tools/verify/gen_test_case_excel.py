#!/usr/bin/env python
# -*- coding:utf-8 -*-

import re
import io
import json
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os
from enum import Enum
import time
import binascii
import struct
import shutil

OUTPUT = "Frigoris"

class Linetype(Enum):
    OTHER = 0
    MODULENAME = 1
    CASENAME  = 2
    DESCRIPTION = 3
    EXPECTATION = 4
    COMMAND = 5
    LABEL = 6
    SKIPLINE = 7

saveDir = os.getcwd()
def getModuleId(core, modulename, id):
    return '0x%02x%02x%02x' % ((binascii.crc32(core.encode('utf-8')) & 0xff, binascii.crc32(modulename.encode('utf-8')) & 0xff, id))

def get_outexcel_name(modulename, filterlist):
    return os.path.join(os.path.relpath(saveDir), OUTPUT + "_" + modulename + time.strftime("%Y%m%d") + ".xlsx")

def get_verification_dir():
    return os.getcwd() + "/"

def mkdir(path):
    if os.path.exists(path):
        shutil.rmtree(path)
    os.makedirs(path)

def check_line_type(line):
    if re.match(r'###',line):
        return Linetype['SKIPLINE']

    if re.match(r'##',line):
        return Linetype['CASENAME']

    if re.match(r'#',line):
        return Linetype['MODULENAME']

    if re.match(r'----',line):
        return Linetype['LABEL']
    
    if line.find("Description") >= 0:
        return Linetype['DESCRIPTION']

    if line.find("Expectation") >= 0:
        return Linetype['EXPECTATION']

    if line.find("FPGA&RTL Params") >= 0:
        return Linetype['COMMAND']

    return Linetype['OTHER']

def pushModule(module, case, filterlist):
    if 'cases' not in module:
        module["cases"] = []

    if case['label'].find("RTL") > 0:
        module["cases"].append(case)
    elif case['label'].find("FPGA") > 0:
        module["cases"].append(case)
    elif case['label'].find("ASIC") > 0:
        module["cases"].append(case)

def markdowntojson(core, filename, filterlist):
    module = {}
    id = 0
    subid = 0
    with io.open(filename, 'r', encoding='UTF-8') as f:
        for line in f.readlines():
            line = line.strip(' \t\n')
            # skip space line and c language structure
            if line == '' or re.match(r'```', line):
                continue
            # print line
            linetype = check_line_type(line)
            # print linetype
            if linetype != Linetype['OTHER']:
                olinetype = linetype

            if olinetype == Linetype['SKIPLINE']:
                continue

            #print("olinetype: %d" % olinetype)
            # skip description
            if linetype == Linetype['DESCRIPTION'] or linetype == Linetype['EXPECTATION'] or linetype == Linetype['COMMAND'] or linetype == Linetype['LABEL']:
                continue
            
            if olinetype == Linetype['MODULENAME']:
                module['name'] = line.strip('#')
                continue

            if olinetype == Linetype['CASENAME']:
                if linetype == Linetype['CASENAME']:
                    if 'case' in dir():
                        # push last case to Module
                        pushModule(module, case, filterlist)
                    # some case do not have command
                    if subid == 0:
                        subid = 1

                    id = id + subid
                    # init new case
                    case = {}
                    subid = 0
                    case['name'] = line.strip('#')
                    case['id'] = getModuleId(core, filename, id)
                continue

            if olinetype == Linetype['LABEL']:
                out_label = []
                case['label'] = line
                continue
            
            if olinetype == Linetype['DESCRIPTION']:
                if 'desc_steps' not in case:
                    case['desc_steps'] = []
                case['desc_steps'].append(line)
                continue
            
            if olinetype == Linetype['EXPECTATION']:
                if 'expect_steps' not in case:
                    case['expect_steps'] = []
                case['expect_steps'].append(line)
                continue

            if olinetype == Linetype['COMMAND']:
                if 'cmd_steps' not in case:
                    case['cmd_steps'] = []

                if 'cmd_comment' not in case:
                    case['cmd_comment'] = []

                subline = line.split("#")
                if len(subline) > 1:
                    case['cmd_comment'].append(subline[1].strip(' '))
                    case['cmd_steps'].append(subline[0].strip(' '))
                else:
                    case['cmd_comment'].append('')
                    case['cmd_steps'].append(line)

                subid = subid + 1
        pushModule(module, case, filterlist)
        #print json.dumps(module, ensure_ascii=False)
        return module

def setColumnWidth(sheet, column, width):
    col_letter = get_column_letter(column)

    if sheet.column_dimensions[col_letter].width > width:
        return

    if width > 100:
        sheet.column_dimensions[col_letter].width = 100
    else:
        sheet.column_dimensions[col_letter].width = width + 6

def setRowHeight(sheet, row, height):
    if height > 100:
        sheet.row_dimensions[row].height = 100
    else:
        sheet.row_dimensions[row].height = height

def excelWriteCell(sheet, row, column, data):
    sheet.cell(row, column, data)
    setColumnWidth(sheet, column, len(data))
    setRowHeight(sheet, row, 15 * (data.count("\n") + 1))

def jsontoexcel(filename, jsondata):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook(filename)
        sheet = wb.create_sheet(u"测试例列表")
        wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(u"测试例列表")
    # if new sheet, write title
    if sheet.max_row == 1:
        #title = [u"测试模块名", u"测试ID号", u"测试项描述", u"测试结果", u"备注"]
        title = [u"测试例", u"测试例ID", u"标签", u"状态", u"测试集", u"模块负责人", u"测试描述", u"测试命令",  u"期望结果", u"测试备注", u"优先级", u"所在部门", u"修改人", u"最后修改时间"]
        for col in range (len (title)):
            excelWriteCell(sheet, 1, col + 1, title[col])

    row = sheet.max_row + 1
    for k in range(len(jsondata['cases'])):
        case = jsondata['cases'][k]
        #print case
        excelWriteCell(sheet, row, 5, jsondata["name"])
        if len(case['cmd_steps']) > 1:
            excelWriteCell(sheet, row, 2, 'MV-' + case["id"] + "~" + 'MV-' + str(hex(int(case['id'], 16) + len(case['cmd_steps']) - 1)))
        else:    
            excelWriteCell(sheet, row, 2, 'MV-' + case["id"])
        excelWriteCell(sheet, row, 1, case["name"])

        label = []
        if 'FPGA' in case["label"]:
            label.append('FPGA')
        if 'RTL' in case["label"]:
            label.append('RTL')
        label.append('ASIC')
        excelWriteCell(sheet, row, 3, ' '.join(label))

        excelWriteCell(sheet, row, 7, '\n'.join(case['desc_steps']))

        excelWriteCell(sheet, row, 8,'\n'.join(case['cmd_steps']))

        excelWriteCell(sheet, row, 9,'\n'.join(case['expect_steps']))

        comment = ''
        for c in range(len(case["cmd_comment"])):
            if case["cmd_comment"][c] != '':
                comment += ("MV-" + str(hex(int(case['id'], 16) + c)) + ": " + case["cmd_comment"][c])
                if c < (len(case["cmd_comment"]) - 1):
                    comment += "\n"
        excelWriteCell(sheet, row, 10, comment)
        row += 1

    wb.save(filename)

def dumpModule(core, mode, modulename, filterlist):
    mkfilename = get_verification_dir() + modulename + '.md'
    print(mkfilename)
    if not os.path.exists(mkfilename):
        print("No this module %s"%modulename)
        return -1
    # in single module
    if mode == 1:
        excelfilename = get_outexcel_name(modulename + '_', filterlist)
        if os.path.exists(excelfilename):
            os.remove(excelfilename)
    else:
        excelfilename = get_outexcel_name("all_", filterlist)
    # markdown to json
    jsondata = markdowntojson(core, mkfilename, filterlist)
    # json to excel
    jsontoexcel(excelfilename, jsondata)

    print("Generate module %s OK"% modulename)
    return 0

def dumpModules(core, filterlist):
    excelfilename = get_outexcel_name('', filterlist)
    if os.path.exists(excelfilename):
        os.remove(excelfilename)

    mddir = get_verification_dir()
    if not os.path.exists(mddir):
        print("No modules found: %s"%mddir)
        return -1
    for root, dirs, files in os.walk(mddir):
        for index in range(len(files)):
            # skip .swp file
            if re.match(r'\.', files[index]):
                continue
            # skip template file
            if files[index] == 'template.md':
                continue
            #print files[index]
            if dumpModule(core, 0, os.path.splitext(files[index])[0], filterlist) < 0:
                return -1

    print("Generate all module OK")
    return 0

def getCommand(command):
    cpuCmds = command.split("(CPU)")
    cpuCmd = cpuCmds[len(cpuCmds) - 1].strip(' \t\n')
    for k in range(len(cpuCmds)):
        cmd = cpuCmds[k].strip(' \t\n')
        if cmd == '':
            continue
        
        sapCmds = cpuCmds[k].split('(SAP)')
        if len(sapCmds) > 1:
            sapCmd = sapCmds[len(sapCmds) - 1].strip(' \t\n')
            if k == (len(cpuCmds) - 1):
                cpuCmd = sapCmds[0].strip(' \t\n')
            return cpuCmd + ';' + sapCmd

        sapCmd = ''

    return cpuCmd
            

if __name__ == "__main__":
    # enter new work path
    if len(sys.argv) != 2:
        print("Command: ")
        print("\tgen_test_case_excel.py modulename \t\tgenerate excel test list for different module")
        print("\tgen_test_case_excel.py all        \t\tgenerate excel test list for all modules")
        sys.exit()
    newDir = os.path.dirname(os.path.abspath(__file__))
    newDir = newDir + "/../../doc/verification/"
    print(newDir)
    os.chdir(newDir)
    arg_core = "c920"
    arg_plat = "RTL"
    
    if sys.argv[1] != "all":
        dumpModule(arg_core, 1, sys.argv[1], arg_plat)
    else:
        dumpModules(arg_core, arg_plat)

    # exit work path
    os.chdir(saveDir)
