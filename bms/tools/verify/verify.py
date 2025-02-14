#!/usr/bin/env python
# -*- coding:utf-8 -*-

import re
import io
import json
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os
from enum import Enum
import time
import binascii
import struct
import shutil

OUTPUT = "Frigoris"

class Linetype(Enum):
    OTHER = 0
    MODULENAME = 1
    CASENAME  = 2
    DESCRIPTION = 3
    EXPECTATION = 4
    COMMAND = 5
    LABEL = 6
    SKIPLINE = 7

load_location = "iram"

saveDir = os.getcwd()
def getModuleId(core, modulename, id):
    return '0x%02x%02x%02x' % ((binascii.crc32(core) & 0xff, binascii.crc32(modulename) & 0xff, id))

def get_outexcel_name(modulename, filterlist):
    return os.path.join(os.path.relpath(saveDir), OUTPUT + "_" + filterlist + "_" + modulename + time.strftime("%Y%m%d") + ".xlsx")

def get_outverify_dir(core, modulename):
    return os.path.join(os.path.relpath(saveDir), OUTPUT + "_" + core + "_RTL_" + modulename  + "_" + load_location + "_" + time.strftime("%Y%m%d"))

def get_verification_dir():
    return "doc/verification/"

def mkdir(path):
    if os.path.exists(path):
        shutil.rmtree(path)
    os.makedirs(path)

def check_line_type(line):
    if re.match(r'###',line):
        return Linetype['SKIPLINE']

    if re.match(r'##',line):
        return Linetype['CASENAME']

    if re.match(r'#',line):
        return Linetype['MODULENAME']

    if re.match(r'----',line):
        return Linetype['LABEL']
    
    if line.find("Description") >= 0:
        return Linetype['DESCRIPTION']

    if line.find("Expectation") >= 0:
        return Linetype['EXPECTATION']

    if line.find("FPGA&RTL Params") >= 0:
        return Linetype['COMMAND']

    return Linetype['OTHER']

def pushModule(module, case, filterlist):
    if 'cases' not in module:
        module["cases"] = []

    if case['label'].find(filterlist) > 0:
        module["cases"].append(case)

def markdowntojson(core, filename, filterlist):
    module = {}
    id = 0
    subid = 0
    with io.open(filename, 'r', encoding='UTF-8') as f:
        for line in f.readlines():
            line = line.strip(' \t\n')
            # skip space line and c language structure
            if line == '' or re.match(r'```', line):
                continue
            # print line
            linetype = check_line_type(line)
            # print linetype
            if linetype != Linetype['OTHER']:
                olinetype = linetype

            if olinetype == Linetype['SKIPLINE']:
                continue

            #print("olinetype: %d" % olinetype)
            # skip description
            if linetype == Linetype['DESCRIPTION'] or linetype == Linetype['EXPECTATION'] or linetype == Linetype['COMMAND'] or linetype == Linetype['LABEL']:
                continue
            
            if olinetype == Linetype['MODULENAME']:
                module['name'] = line.strip('#')
                continue

            if olinetype == Linetype['CASENAME']:
                if linetype == Linetype['CASENAME']:
                    if 'case' in dir():
                        # push last case to Module
                        pushModule(module, case, filterlist)
                    # some case do not have command
                    if subid == 0:
                        subid = 1

                    id = id + subid
                    # init new case
                    case = {}
                    subid = 0
                    case['name'] = line.strip('#')
                    case['id'] = getModuleId(core, filename, id)
                continue

            if olinetype == Linetype['LABEL']:
                case['label'] = line
                continue
            
            if olinetype == Linetype['DESCRIPTION']:
                if 'desc_steps' not in case:
                    case['desc_steps'] = []
                case['desc_steps'].append(line)
                continue
            
            if olinetype == Linetype['EXPECTATION']:
                if 'expect_steps' not in case:
                    case['expect_steps'] = []
                case['expect_steps'].append(line)
                continue

            if olinetype == Linetype['COMMAND']:
                if 'cmd_steps' not in case:
                    case['cmd_steps'] = []

                if 'cmd_comment' not in case:
                    case['cmd_comment'] = []

                subline = line.split("#")
                if len(subline) > 1:
                    case['cmd_comment'].append(subline[1].strip(' '))
                    case['cmd_steps'].append(subline[0].strip(' '))
                else:
                    case['cmd_comment'].append('')
                    case['cmd_steps'].append(line)

                subid = subid + 1
        pushModule(module, case, filterlist)
        #print json.dumps(module, ensure_ascii=False)
        return module

def setColumnWidth(sheet, column, width):
    col_letter = get_column_letter(column)

    if sheet.column_dimensions[col_letter].width > width:
        return

    if width > 100:
        sheet.column_dimensions[col_letter].width = 100
    else:
        sheet.column_dimensions[col_letter].width = width + 6

def setRowHeight(sheet, row, height):
    if height > 100:
        sheet.row_dimensions[row].height = 100
    else:
        sheet.row_dimensions[row].height = height

def excelWriteCell(sheet, row, column, data):
    sheet.cell(row, column, data)
    setColumnWidth(sheet, column, len(data))
    setRowHeight(sheet, row, 15 * (data.count("\n") + 1))

def jsontoexcel(filename, jsondata):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook(filename)
        sheet = wb.create_sheet(u"测试列表")
        wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(u"测试列表")
    # if new sheet, write title
    if sheet.max_row == 1:
        title = [u"测试模块名", u"测试ID号", u"测试项描述", u"测试结果", u"备注"]
        for col in range (len (title)):
            excelWriteCell(sheet, 1, col + 1, title[col])

    row = sheet.max_row + 1
    for k in range(len(jsondata['cases'])):
        case = jsondata['cases'][k]
        #print case
        excelWriteCell(sheet, row, 1, jsondata["name"])
        if len(case['cmd_steps']) > 1:
            excelWriteCell(sheet, row, 2, 'MV-' + case["id"] + "~" + 'MV-' + str(hex(int(case['id'], 16) + len(case['cmd_steps']) - 1)))
        else:    
            excelWriteCell(sheet, row, 2, 'MV-' + case["id"])
        excelWriteCell(sheet, row, 3, case["name"])
        comment = ''
        for c in range(len(case["cmd_comment"])):
            if case["cmd_comment"][c] != '':
                comment += ("MV-" + str(hex(int(case['id'], 16) + c)) + ": " + case["cmd_comment"][c])
                if c < (len(case["cmd_comment"]) - 1):
                    comment += "\n"
        excelWriteCell(sheet, row, 5, comment)
        row += 1

    wb.save(filename)

def dumpModule(core, mode, modulename, filterlist):
    mkfilename = get_verification_dir() + modulename + '.md'
    if not os.path.exists(mkfilename):
        print("No this module %s"%modulename)
        return -1
    # in single module
    if mode == 1:
        excelfilename = get_outexcel_name(core + '_' + modulename + '_', filterlist)
        if os.path.exists(excelfilename):
            os.remove(excelfilename)
    else:
        excelfilename = get_outexcel_name(core, filterlist)
    # markdown to json
    jsondata = markdowntojson(core, mkfilename, filterlist)
    # json to excel
    jsontoexcel(excelfilename, jsondata)

    print("Generate module %s OK"% modulename)
    return 0

def dumpModules(core, filterlist):
    excelfilename = get_outexcel_name('', filterlist)
    if os.path.exists(excelfilename):
        os.remove(excelfilename)

    mddir = get_verification_dir()
    if not os.path.exists(mddir):
        print("No modules found: %s"%mddir)
        return -1
    for root, dirs, files in os.walk(mddir):
        for index in range(len(files)):
            # skip .swp file
            if re.match(r'\.', files[index]):
                continue
            # skip template file
            if files[index] == 'template.md':
                continue
            #print files[index]
            if dumpModule(core, 0, os.path.splitext(files[index])[0], filterlist) < 0:
                return -1

    print("Generate all module OK")
    return 0

'''
struct rtl_descs {
    int magic;
    int id;
}

struct rtl_case {
    int id;
    int len;
    char cmd[8];
}
'''
DESC_MAGIC = 0x97979897

def getCommand(command):
    cpuCmds = command.split("(CPU)")
    cpuCmd = cpuCmds[len(cpuCmds) - 1].strip(' \t\n')
    for k in range(len(cpuCmds)):
        cmd = cpuCmds[k].strip(' \t\n')
        if cmd == '':
            continue
        
        sapCmds = cpuCmds[k].split('(SAP)')
        if len(sapCmds) > 1:
            sapCmd = sapCmds[len(sapCmds) - 1].strip(' \t\n')
            if k == (len(cpuCmds) - 1):
                cpuCmd = sapCmds[0].strip(' \t\n')
            return cpuCmd + ';' + sapCmd

        sapCmd = ''

    return cpuCmd
            



def generateRTLCase(rootdir, id, command):
    commands = command.split(';')
    # mkdir case dir
    casedir = rootdir + '/MV-' + hex(id)
    mkdir(casedir)
    filename = casedir + '/' + 'rtl_case.bin'
    with io.open(filename, 'wb') as f:
        for k in range(len(commands)):
            string = getCommand(commands[k].strip(' '))
            strlen = len(string)

            # for situation last ; or ;; 
            if strlen == 0:
                continue

            caselen = (strlen + 8) & ~0x7
            # print string
            soffs = string.find(';')
            if k == 0:
                data = struct.pack("I4i{}s".format(caselen), DESC_MAGIC, id, id, strlen, soffs, str(string))
            else:
                data = struct.pack("3i{}s".format(caselen), id, strlen, soffs, str(string))
            f.write(data)
        data = struct.pack("3i", 0, -1, 0)
        f.write(data)
        f.close()
    
    print("Generate MV-%s rtl_case.bin"%hex(id))
    print("Cmd: %s"%str(command))
    return

def generateModuleRTLCases(core, modulename, filterlist):
    mkfilename = get_verification_dir() + modulename + '.md'
    if not os.path.exists(mkfilename):
        print("No this module %s"%modulename)
        return -1

    # markdown to json
    jsondata = markdowntojson(core, mkfilename, filterlist)
    # mkdir rootdir
    rootdir = get_outverify_dir(core, modulename)
    mkdir(rootdir)

    for k in range(len(jsondata['cases'])):
        case = jsondata['cases'][k]
        if 'cmd_steps' not in case:
            generateRTLCase(rootdir, int(case['id'], 16) + h, '')
            continue

        for h in range(len(case['cmd_steps'])):
            generateRTLCase(rootdir, int(case['id'], 16) + h, case['cmd_steps'][h])

    print("Generate module %s all case OK"%modulename)

def buildImage(config, location):
    os.system("make clean")
    print "make " + config
    os.system("make " + config)
    if config == 'e906_rtl_rom_defconfig':
        if location == 'ddr':
            print("modify out/.config in ddr mode\n")
            os.system("sed -i \"s/# CONFIG_TB_RUN_DDR is not set/CONFIG_TB_RUN_DDR=y/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_DDR1=y/# CONFIG_TB_RUN_DDR1 is not set/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_IRAM=y/# CONFIG_TB_RUN_IRAM is not set/\" out/.config")
        elif location == 'ddr1':
            print("modify out/.config in ddr1 mode\n")
            os.system("sed -i \"s/CONFIG_TB_RUN_DDR=y/# CONFIG_TB_RUN_DDR is not set/\" out/.config")
            os.system("sed -i \"s/# CONFIG_TB_RUN_DDR1 is not set/CONFIG_TB_RUN_DDR1=y/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_IRAM=y/# CONFIG_TB_RUN_IRAM is not set/\" out/.config")
            
    elif config == 'e906_rtl_orom_defconfig':
        if location == 'ddr':
            print("modify out/.config in ddr mode\n")
            os.system("sed -i \"s/# CONFIG_TB is not set/CONFIG_TB=y/\" out/.config")
            os.system("sed -i \"s/# CONFIG_TB_RUN_DDR is not set/CONFIG_TB_RUN_DDR=y/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_DDR1=y/# CONFIG_TB_RUN_DDR1 is not set/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_IRAM=y/# CONFIG_TB_RUN_IRAM is not set/\" out/.config")
        elif location == 'ddr1':
            print("modify out/.config in ddr1 mode\n")
            #os.system("sed -i \"s/# CONFIG_TB is not set/CONFIG_TB=y/\" out/.config")
            os.system("sed -i \"s/# CONFIG_TB_RUN_DDR1 is not set/CONFIG_TB_RUN_DDR1=y/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_DDR=y/# CONFIG_TB_RUN_DDR is not set/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_IRAM=y/# CONFIG_TB_RUN_IRAM is not set/\" out/.config")
    elif config == 'c920_rtl_ddr_soc_defconfig':
        if location == 'ddr':
            print("modify out/.config in ddr mode\n")
            os.system("sed -i \"s/# CONFIG_TB_RUN_DDR is not set/CONFIG_TB_RUN_DDR=y/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_DDR1=y/# CONFIG_TB_RUN_DDR1 is not set/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_IRAM=y/# CONFIG_TB_RUN_IRAM is not set/\" out/.config")
        elif location == 'ddr1':
            print("modify out/.config in ddr1 mode\n")
            os.system("sed -i \"s/# CONFIG_TB_RUN_DDR1 is not set/CONFIG_TB_RUN_DDR1=y/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_DDR=y/# CONFIG_TB_RUN_DDR is not set/\" out/.config")
            os.system("sed -i \"s/CONFIG_TB_RUN_IRAM=y/# CONFIG_TB_RUN_IRAM is not set/\" out/.config")
    os.system("make")

def moveImagetoDst(core, modulename):
    rootdir = get_outverify_dir(core, modulename)
    for root, dirs, files in os.walk(rootdir):
        for k in range(len(dirs)):
            if os.path.exists("out/soc.bin"):
                os.system("cp " + "out/soc.bin " + os.path.join(rootdir + '/', dirs[k]))
                os.system("cp " + "out/soc.S " + os.path.join(rootdir + '/', dirs[k]))
            if os.path.exists("out/rom.bin"):
                if os.path.exists("out/rom_data.bin")==False:
                    os.system("./tools/verify/gen_romdata.sh")
                os.system("cp " + "out/rom_data.bin " + os.path.join(rootdir + '/', dirs[k]))
                os.system("cp " + "out/rom.bin " + os.path.join(rootdir + '/', dirs[k]))
                os.system("cp " + "out/rom.S " + os.path.join(rootdir + '/', dirs[k]))

def generateRTLbin(core, modulename, location):
    # norom = ['irq', 'timer', 'i2c', 'wdt', 'cuart', 'sysm']
    # if modulename in norom:
    #     config = core + '_rtl_norom_defconfig'
    # else:
    #     config = core + '_rtl_defconfig'
    if core == 'e906':
        config = core + '_rtl_rom_defconfig'
    else:
        config = 'e906_rtl_orom_defconfig'

    buildImage(config, location)
    moveImagetoDst(core, modulename)
    
    if location == 'iram':
        config = core + '_rtl_iram_soc_defconfig'
    else:
        config = core + '_rtl_ddr_soc_defconfig'

    buildImage(config, location)
    moveImagetoDst(core, modulename)

    os.system("tar -czf " + get_outverify_dir(core, modulename) + ".tar.gz " + "-C " + os.path.dirname(get_outverify_dir(core, modulename)) + " " + os.path.basename(get_outverify_dir(core, modulename)))
    os.system("rm -rf " + get_outverify_dir(core, modulename))
    

if __name__ == "__main__":
    if sys.argv[1] == 'rtlcases':
        if len(sys.argv) != 5:
            print("Command: ")
            print("\tverify.py cpucore modulename platform\t\t\tgenerate excel test list for different platforms(RTL or FPGA)")
            print("\tverify.py rtlcases cpucore modulename location\t\tgenerate module case for rtl test in different location(ddr or iram)")
            sys.exit()
    else:
        if len(sys.argv) != 4:
            print("Command: ")
            print("\tverify.py cpucore modulename platform\t\t\tgenerate excel test list for different platforms(RTL or FPGA)")
            print("\tverify.py rtlcases cpucore modulename location\t\tgenerate module case for rtl test in different location(ddr or iram)")
            sys.exit()

    # enter new work path
    newDir = os.path.dirname(__file__)
    os.chdir(newDir)
    os.chdir('../../')
    if sys.argv[1] == 'rtlcases':
        # dump cmd
        load_location = sys.argv[4]
        generateModuleRTLCases(sys.argv[2], sys.argv[3], "RTL")
        generateRTLbin(sys.argv[2], sys.argv[3], sys.argv[4])
    else:
        if sys.argv[2] != "all":
            dumpModule(sys.argv[1], 1, sys.argv[2], sys.argv[3])
        else:
            dumpModules(sys.argv[1], sys.argv[3])
    # exit work path
    os.chdir(saveDir)
